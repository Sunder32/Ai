"""
Сервис экспорта конфигураций ПК в различные форматы
PDF, Excel, CSV
"""
import io
import csv
from decimal import Decimal
from typing import Optional, Dict, Any
from datetime import datetime

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ConfigurationExportService:
    """Сервис экспорта конфигураций"""
    
    # Ссылки на магазины (шаблоны поиска)
    SHOP_LINKS = {
        'dns': 'https://www.dns-shop.ru/search/?q=',
        'citilink': 'https://www.citilink.ru/search/?text=',
        'mvideo': 'https://www.mvideo.ru/product-list-page?q=',
        'regard': 'https://www.regard.ru/catalog?search=',
    }
    
    def __init__(self, configuration):
        self.config = configuration
        self.workspace = getattr(configuration, 'workspace_setups', None)
        if self.workspace:
            self.workspace = self.workspace.first()
    
    def get_components_list(self) -> list:
        """Получить список компонентов конфигурации"""
        components = []
        
        component_map = [
            ('Процессор', self.config.cpu),
            ('Видеокарта', self.config.gpu),
            ('Материнская плата', self.config.motherboard),
            ('Оперативная память', self.config.ram),
            ('Накопитель (основной)', self.config.storage_primary),
            ('Накопитель (доп.)', self.config.storage_secondary),
            ('Блок питания', self.config.psu),
            ('Корпус', self.config.case),
            ('Охлаждение', self.config.cooling),
        ]
        
        for name, component in component_map:
            if component:
                components.append({
                    'type': name,
                    'name': str(component),
                    'price': float(component.price) if component.price else 0,
                    'search_query': component.name,
                })
        
        return components
    
    def get_peripherals_list(self) -> list:
        """Получить список периферии"""
        if not self.workspace:
            return []
        
        peripherals = []
        peripheral_map = [
            ('Монитор (основной)', self.workspace.monitor_primary),
            ('Монитор (доп.)', self.workspace.monitor_secondary),
            ('Клавиатура', self.workspace.keyboard),
            ('Мышь', self.workspace.mouse),
            ('Наушники', self.workspace.headset),
            ('Веб-камера', self.workspace.webcam),
            ('Микрофон', self.workspace.microphone),
            ('Стол', self.workspace.desk),
            ('Кресло', self.workspace.chair),
        ]
        
        for name, peripheral in peripheral_map:
            if peripheral:
                peripherals.append({
                    'type': name,
                    'name': str(peripheral),
                    'price': float(peripheral.price) if peripheral.price else 0,
                    'search_query': peripheral.name if hasattr(peripheral, 'name') else str(peripheral),
                })
        
        return peripherals
    
    def generate_shop_links(self, component_name: str) -> Dict[str, str]:
        """Генерация ссылок на магазины для компонента"""
        import urllib.parse
        encoded_name = urllib.parse.quote(component_name)
        return {
            shop: f"{url}{encoded_name}"
            for shop, url in self.SHOP_LINKS.items()
        }
    
    def export_to_csv(self) -> io.StringIO:
        """Экспорт в CSV формат"""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Заголовок
        writer.writerow(['Конфигурация ПК:', self.config.name])
        writer.writerow(['Дата создания:', self.config.created_at.strftime('%d.%m.%Y')])
        writer.writerow([])
        
        # Компоненты ПК
        writer.writerow(['=== Компоненты ПК ==='])
        writer.writerow(['Тип', 'Название', 'Цена (₽)', 'DNS', 'Citilink'])
        
        components = self.get_components_list()
        for comp in components:
            links = self.generate_shop_links(comp['search_query'])
            writer.writerow([
                comp['type'],
                comp['name'],
                f"{comp['price']:.2f}",
                links['dns'],
                links['citilink'],
            ])
        
        # Итого ПК
        pc_total = sum(c['price'] for c in components)
        writer.writerow([])
        writer.writerow(['Итого ПК:', '', f"{pc_total:.2f}"])
        
        # Периферия
        peripherals = self.get_peripherals_list()
        if peripherals:
            writer.writerow([])
            writer.writerow(['=== Периферия ==='])
            writer.writerow(['Тип', 'Название', 'Цена (₽)', 'DNS', 'Citilink'])
            
            for periph in peripherals:
                links = self.generate_shop_links(periph['search_query'])
                writer.writerow([
                    periph['type'],
                    periph['name'],
                    f"{periph['price']:.2f}",
                    links['dns'],
                    links['citilink'],
                ])
            
            periph_total = sum(p['price'] for p in peripherals)
            writer.writerow([])
            writer.writerow(['Итого периферия:', '', f"{periph_total:.2f}"])
        
        # Общий итог
        total = float(self.config.total_price) if self.config.total_price else pc_total
        if self.workspace and self.workspace.total_price:
            total = float(self.workspace.total_price)
        
        writer.writerow([])
        writer.writerow(['=== ОБЩИЙ ИТОГ ===', '', f"{total:.2f}"])
        
        output.seek(0)
        return output
    
    def export_to_excel(self) -> Optional[io.BytesIO]:
        """Экспорт в Excel формат"""
        if not OPENPYXL_AVAILABLE:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Конфигурация ПК"
        
        # Стили
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        currency_format = '#,##0.00 ₽'
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        
        # Заголовок
        ws['A1'] = f"Конфигурация: {self.config.name}"
        ws['A1'].font = header_font
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Дата: {self.config.created_at.strftime('%d.%m.%Y')}"
        ws['A2'].font = subheader_font
        
        # Компоненты ПК
        row = 4
        ws[f'A{row}'] = "Компоненты ПК"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font_white
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        headers = ['Тип', 'Название', 'Цена', 'DNS', 'Citilink']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        components = self.get_components_list()
        for comp in components:
            row += 1
            links = self.generate_shop_links(comp['search_query'])
            ws.cell(row=row, column=1, value=comp['type'])
            ws.cell(row=row, column=2, value=comp['name'])
            price_cell = ws.cell(row=row, column=3, value=comp['price'])
            price_cell.number_format = currency_format
            ws.cell(row=row, column=4, value=links['dns'])
            ws.cell(row=row, column=5, value=links['citilink'])
        
        # Итого ПК
        row += 1
        pc_total = sum(c['price'] for c in components)
        ws.cell(row=row, column=2, value="Итого ПК:").font = Font(bold=True)
        total_cell = ws.cell(row=row, column=3, value=pc_total)
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True)
        
        # Периферия
        peripherals = self.get_peripherals_list()
        if peripherals:
            row += 2
            ws[f'A{row}'] = "Периферия"
            ws[f'A{row}'].font = header_font_white
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 1
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            for periph in peripherals:
                row += 1
                links = self.generate_shop_links(periph['search_query'])
                ws.cell(row=row, column=1, value=periph['type'])
                ws.cell(row=row, column=2, value=periph['name'])
                price_cell = ws.cell(row=row, column=3, value=periph['price'])
                price_cell.number_format = currency_format
                ws.cell(row=row, column=4, value=links['dns'])
                ws.cell(row=row, column=5, value=links['citilink'])
            
            row += 1
            periph_total = sum(p['price'] for p in peripherals)
            ws.cell(row=row, column=2, value="Итого периферия:").font = Font(bold=True)
            total_cell = ws.cell(row=row, column=3, value=periph_total)
            total_cell.number_format = currency_format
            total_cell.font = Font(bold=True)
        
        # Общий итог
        row += 2
        total = float(self.config.total_price) if self.config.total_price else pc_total
        ws.cell(row=row, column=2, value="ОБЩИЙ ИТОГ:").font = Font(bold=True, size=12)
        total_cell = ws.cell(row=row, column=3, value=total)
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True, size=12)
        total_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Автоширина колонок
        for col in range(1, 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_to_pdf(self) -> Optional[io.BytesIO]:
        """Экспорт в PDF формат"""
        if not REPORTLAB_AVAILABLE:
            return None
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, 
                               rightMargin=20*mm, leftMargin=20*mm,
                               topMargin=20*mm, bottomMargin=20*mm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Заголовок
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )
        elements.append(Paragraph(f"Конфигурация: {self.config.name}", title_style))
        elements.append(Paragraph(f"Дата: {self.config.created_at.strftime('%d.%m.%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Таблица компонентов
        elements.append(Paragraph("Компоненты ПК", styles['Heading2']))
        
        data = [['Тип', 'Название', 'Цена (руб.)']]
        components = self.get_components_list()
        for comp in components:
            data.append([comp['type'], comp['name'][:40], f"{comp['price']:,.0f}"])
        
        pc_total = sum(c['price'] for c in components)
        data.append(['', 'Итого ПК:', f"{pc_total:,.0f}"])
        
        table = Table(data, colWidths=[80, 250, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Периферия
        peripherals = self.get_peripherals_list()
        if peripherals:
            elements.append(Paragraph("Периферия", styles['Heading2']))
            
            data = [['Тип', 'Название', 'Цена (руб.)']]
            for periph in peripherals:
                data.append([periph['type'], periph['name'][:40], f"{periph['price']:,.0f}"])
            
            periph_total = sum(p['price'] for p in peripherals)
            data.append(['', 'Итого периферия:', f"{periph_total:,.0f}"])
            
            table = Table(data, colWidths=[80, 250, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Общий итог
        total = float(self.config.total_price) if self.config.total_price else pc_total
        total_style = ParagraphStyle(
            'Total',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#006400')
        )
        elements.append(Paragraph(f"ОБЩИЙ ИТОГ: {total:,.0f} руб.", total_style))
        
        doc.build(elements)
        output.seek(0)
        return output


class PowerCalculatorService:
    """Калькулятор энергопотребления системы"""
    
    # Средняя цена электричества в России (руб/кВт*ч)
    ELECTRICITY_PRICE = 5.5
    
    # Дополнительное потребление (вентиляторы, подсветка и т.д.)
    MISC_POWER = 30  # Вт
    
    def __init__(self, configuration):
        self.config = configuration
    
    def calculate_system_tdp(self) -> Dict[str, Any]:
        """Рассчитать TDP всей системы"""
        components_tdp = {}
        total_tdp = 0
        
        # CPU TDP
        if self.config.cpu and hasattr(self.config.cpu, 'tdp'):
            cpu_tdp = self.config.cpu.tdp
            components_tdp['cpu'] = {
                'name': str(self.config.cpu),
                'tdp': cpu_tdp,
                'peak_tdp': int(cpu_tdp * 1.3)  # Пиковое потребление +30%
            }
            total_tdp += cpu_tdp
        
        # GPU TDP
        if self.config.gpu and hasattr(self.config.gpu, 'tdp'):
            gpu_tdp = self.config.gpu.tdp
            components_tdp['gpu'] = {
                'name': str(self.config.gpu),
                'tdp': gpu_tdp,
                'peak_tdp': int(gpu_tdp * 1.2),  # Пиковое потребление +20%
                'recommended_psu': self.config.gpu.recommended_psu if hasattr(self.config.gpu, 'recommended_psu') else None
            }
            total_tdp += gpu_tdp
        
        # Материнская плата (~50W)
        if self.config.motherboard:
            mb_tdp = 50
            components_tdp['motherboard'] = {
                'name': str(self.config.motherboard),
                'tdp': mb_tdp
            }
            total_tdp += mb_tdp
        
        # RAM (~5W per module)
        if self.config.ram:
            modules = self.config.ram.modules if hasattr(self.config.ram, 'modules') else 2
            ram_tdp = modules * 5
            components_tdp['ram'] = {
                'name': str(self.config.ram),
                'tdp': ram_tdp,
                'modules': modules
            }
            total_tdp += ram_tdp
        
        # Storage (SSD ~5W, HDD ~10W)
        storage_tdp = 0
        if self.config.storage_primary:
            st_type = getattr(self.config.storage_primary, 'storage_type', 'ssd_nvme')
            st_power = 10 if 'hdd' in st_type.lower() else 5
            storage_tdp += st_power
        if self.config.storage_secondary:
            st_type = getattr(self.config.storage_secondary, 'storage_type', 'hdd')
            st_power = 10 if 'hdd' in st_type.lower() else 5
            storage_tdp += st_power
        
        if storage_tdp > 0:
            components_tdp['storage'] = {'tdp': storage_tdp}
            total_tdp += storage_tdp
        
        # Cooling (fans, pumps)
        if self.config.cooling:
            cooling_type = getattr(self.config.cooling, 'cooling_type', 'air')
            cooling_tdp = 15 if cooling_type == 'air' else 25
            components_tdp['cooling'] = {
                'name': str(self.config.cooling),
                'tdp': cooling_tdp
            }
            total_tdp += cooling_tdp
        
        # Misc (case fans, RGB)
        total_tdp += self.MISC_POWER
        components_tdp['misc'] = {'tdp': self.MISC_POWER}
        
        return {
            'components': components_tdp,
            'total_tdp': total_tdp,
            'peak_tdp': int(total_tdp * 1.3)
        }
    
    def recommend_psu(self) -> Dict[str, Any]:
        """Рекомендация блока питания"""
        tdp_data = self.calculate_system_tdp()
        total_tdp = tdp_data['total_tdp']
        peak_tdp = tdp_data['peak_tdp']
        
        # Рекомендуемая мощность с запасом 20-30%
        recommended_20 = int(total_tdp * 1.2)
        recommended_30 = int(total_tdp * 1.3)
        
        # Округляем до стандартных значений БП
        standard_wattages = [450, 500, 550, 600, 650, 700, 750, 800, 850, 1000, 1200, 1500]
        
        min_psu = next((w for w in standard_wattages if w >= recommended_20), 1500)
        optimal_psu = next((w for w in standard_wattages if w >= recommended_30), 1500)
        
        # Проверяем текущий БП
        current_psu = None
        psu_sufficient = None
        if self.config.psu:
            current_psu = {
                'name': str(self.config.psu),
                'wattage': self.config.psu.wattage
            }
            psu_sufficient = self.config.psu.wattage >= min_psu
        
        return {
            'system_tdp': total_tdp,
            'peak_tdp': peak_tdp,
            'recommended_minimum': min_psu,
            'recommended_optimal': optimal_psu,
            'current_psu': current_psu,
            'current_sufficient': psu_sufficient,
            'headroom_percent': round((self.config.psu.wattage - total_tdp) / total_tdp * 100, 1) if current_psu else None
        }
    
    def calculate_electricity_cost(self, hours_per_day: int = 8, load_percent: float = 0.7) -> Dict[str, Any]:
        """Расчёт стоимости электричества"""
        tdp_data = self.calculate_system_tdp()
        total_tdp = tdp_data['total_tdp']
        
        # Средняя нагрузка
        avg_consumption_watts = total_tdp * load_percent
        
        # Потребление в кВт*ч
        daily_kwh = (avg_consumption_watts * hours_per_day) / 1000
        monthly_kwh = daily_kwh * 30
        yearly_kwh = daily_kwh * 365
        
        # Стоимость
        daily_cost = daily_kwh * self.ELECTRICITY_PRICE
        monthly_cost = monthly_kwh * self.ELECTRICITY_PRICE
        yearly_cost = yearly_kwh * self.ELECTRICITY_PRICE
        
        return {
            'parameters': {
                'hours_per_day': hours_per_day,
                'load_percent': load_percent * 100,
                'electricity_price': self.ELECTRICITY_PRICE
            },
            'consumption': {
                'avg_watts': round(avg_consumption_watts, 1),
                'daily_kwh': round(daily_kwh, 2),
                'monthly_kwh': round(monthly_kwh, 2),
                'yearly_kwh': round(yearly_kwh, 2)
            },
            'cost_rub': {
                'daily': round(daily_cost, 2),
                'monthly': round(monthly_cost, 2),
                'yearly': round(yearly_cost, 2)
            }
        }


class CompatibilityChecker:
    """Расширенная проверка совместимости компонентов"""
    
    def __init__(self, configuration):
        self.config = configuration
        self.issues = []
        self.warnings = []
    
    def check_all(self) -> Dict[str, Any]:
        """Выполнить все проверки совместимости"""
        self.issues = []
        self.warnings = []
        
        self._check_cpu_socket()
        self._check_ram_type()
        self._check_gpu_length()
        self._check_cooler_tdp()
        self._check_psu_power()
        self._check_case_form_factor()
        
        return {
            'compatible': len(self.issues) == 0,
            'issues': self.issues,
            'warnings': self.warnings,
            'checks_passed': 6 - len(self.issues)
        }
    
    def _check_cpu_socket(self):
        """Проверка совместимости сокета CPU и материнской платы"""
        if self.config.cpu and self.config.motherboard:
            cpu_socket = self.config.cpu.socket
            mb_socket = self.config.motherboard.socket
            
            if cpu_socket != mb_socket:
                self.issues.append({
                    'type': 'socket_mismatch',
                    'severity': 'critical',
                    'message': f'Несовместимость сокетов: CPU ({cpu_socket}) ≠ Материнская плата ({mb_socket})',
                    'components': ['cpu', 'motherboard']
                })
    
    def _check_ram_type(self):
        """Проверка совместимости типа RAM и материнской платы"""
        if self.config.ram and self.config.motherboard:
            ram_type = self.config.ram.memory_type
            mb_ram_type = self.config.motherboard.memory_type
            
            if ram_type != mb_ram_type:
                self.issues.append({
                    'type': 'ram_type_mismatch',
                    'severity': 'critical',
                    'message': f'Несовместимость типа памяти: RAM ({ram_type}) ≠ Материнская плата ({mb_ram_type})',
                    'components': ['ram', 'motherboard']
                })
            
            # Проверка количества модулей
            ram_modules = getattr(self.config.ram, 'modules', 2)
            mb_slots = self.config.motherboard.memory_slots
            
            if ram_modules > mb_slots:
                self.issues.append({
                    'type': 'ram_slots_exceeded',
                    'severity': 'critical',
                    'message': f'Недостаточно слотов RAM: требуется {ram_modules}, доступно {mb_slots}',
                    'components': ['ram', 'motherboard']
                })
    
    def _check_gpu_length(self):
        """Проверка длины GPU vs корпус"""
        if self.config.gpu and self.config.case:
            max_gpu_length = getattr(self.config.case, 'max_gpu_length', None)
            
            # Оценочная длина GPU по модели (если нет данных)
            gpu_length = 300  # По умолчанию 300мм для топовых карт
            
            if max_gpu_length and gpu_length > max_gpu_length:
                self.issues.append({
                    'type': 'gpu_too_long',
                    'severity': 'critical',
                    'message': f'Видеокарта может не поместиться в корпус (макс. {max_gpu_length}мм)',
                    'components': ['gpu', 'case']
                })
            elif max_gpu_length is None:
                self.warnings.append({
                    'type': 'gpu_length_unknown',
                    'message': 'Не указана максимальная длина GPU для корпуса. Проверьте совместимость вручную.',
                    'components': ['gpu', 'case']
                })
    
    def _check_cooler_tdp(self):
        """Проверка TDP охлаждения vs TDP процессора"""
        if self.config.cpu and self.config.cooling:
            cpu_tdp = self.config.cpu.tdp
            cooler_max_tdp = getattr(self.config.cooling, 'max_tdp', None)
            
            if cooler_max_tdp:
                if cooler_max_tdp < cpu_tdp:
                    self.issues.append({
                        'type': 'cooler_insufficient',
                        'severity': 'warning',
                        'message': f'Охлаждение может быть недостаточным: кулер до {cooler_max_tdp}Вт, CPU {cpu_tdp}Вт',
                        'components': ['cpu', 'cooling']
                    })
                elif cooler_max_tdp < cpu_tdp * 1.2:
                    self.warnings.append({
                        'type': 'cooler_marginal',
                        'message': f'Запас по охлаждению небольшой. Рекомендуется кулер с TDP >{cpu_tdp * 1.2:.0f}Вт',
                        'components': ['cpu', 'cooling']
                    })
            
            # Проверка совместимости сокетов охлаждения
            socket_compat = getattr(self.config.cooling, 'socket_compatibility', '')
            cpu_socket = self.config.cpu.socket
            
            if socket_compat and cpu_socket not in socket_compat:
                self.issues.append({
                    'type': 'cooler_socket_mismatch',
                    'severity': 'critical',
                    'message': f'Кулер не поддерживает сокет {cpu_socket}',
                    'components': ['cpu', 'cooling']
                })
    
    def _check_psu_power(self):
        """Проверка мощности БП"""
        if self.config.psu:
            psu_wattage = self.config.psu.wattage
            
            # Рассчитываем требуемую мощность
            required_power = 100  # Базовое потребление
            
            if self.config.cpu:
                required_power += self.config.cpu.tdp
            
            if self.config.gpu:
                gpu_recommended = getattr(self.config.gpu, 'recommended_psu', None)
                if gpu_recommended:
                    required_power = max(required_power, gpu_recommended * 0.8)
                else:
                    required_power += self.config.gpu.tdp
            
            # Добавляем 20% запас
            required_power *= 1.2
            
            if psu_wattage < required_power:
                self.issues.append({
                    'type': 'psu_insufficient',
                    'severity': 'warning',
                    'message': f'БП может быть недостаточно: {psu_wattage}Вт, рекомендуется ≥{required_power:.0f}Вт',
                    'components': ['psu']
                })
    
    def _check_case_form_factor(self):
        """Проверка форм-фактора корпуса и материнской платы"""
        if self.config.motherboard and self.config.case:
            mb_ff = self.config.motherboard.form_factor.upper()
            case_ff = self.config.case.form_factor.upper()
            
            # Совместимость форм-факторов
            compatibility_map = {
                'ATX': ['ATX', 'FULL TOWER', 'MID TOWER'],
                'MICRO-ATX': ['ATX', 'MICRO-ATX', 'FULL TOWER', 'MID TOWER', 'MINI TOWER'],
                'MINI-ITX': ['ATX', 'MICRO-ATX', 'MINI-ITX', 'FULL TOWER', 'MID TOWER', 'MINI TOWER', 'SFF'],
            }
            
            compatible_cases = compatibility_map.get(mb_ff, [mb_ff])
            
            if case_ff not in compatible_cases and mb_ff not in case_ff:
                self.warnings.append({
                    'type': 'form_factor_check',
                    'message': f'Проверьте совместимость форм-факторов: материнка {mb_ff}, корпус {case_ff}',
                    'components': ['motherboard', 'case']
                })
